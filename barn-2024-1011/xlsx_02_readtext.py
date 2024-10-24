import openpyxl
def print_all_sheet_contents(file_path):
    wb = openpyxl.load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"워크시트: {sheet}")
        for row in ws.iter_rows(values_only=True):
            print([str(cell) if cell is not None else '' for cell in row])
        print('-' * 40)  # 각 워크시트 구분선
file_path = 'example.xlsx'
print_all_sheet_contents(file_path)
