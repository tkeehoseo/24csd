import sys
import openpyxl

def extract_text_from_xlsx(input_xlsx, output_txt):
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(input_xlsx)
    
    # 출력할 텍스트 파일 열기
    with open(output_txt, 'w', encoding='utf-8') as f:
        # 모든 워크시트를 순회
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f.write(f"--- Worksheet: {sheet_name} ---\n")
            
            # 워크시트의 모든 행을 순회하며 텍스트 추출
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                f.write("\t".join(row_data) + "\n")
            
            f.write("\n")  # 워크시트 간 구분을 위한 빈 줄

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("사용법: python scxlsx2txt.py <input_xlsx> <output_txt>")
    else:
        input_xlsx = sys.argv[1]
        output_txt = sys.argv[2]
        extract_text_from_xlsx(input_xlsx, output_txt)
        print(f"텍스트가 {output_txt} 파일에 저장되었습니다.")
